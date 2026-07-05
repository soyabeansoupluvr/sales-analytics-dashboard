"""M9 — Ingestion & Validation.

Accepts CSV or XLSX uploads, validates the expected schema, rejects workbook
macros and formulas, screens text values for spreadsheet formula injection,
and emits a data-quality report for M10 Cleaning / ETL.

Security alignment:

* OWASP Top 10:2025 A05 Injection, where spreadsheet formula injection applies.
* NIST SSDF PW.5.1 secure coding practices for validating inputs and encoding
  outputs. Output modules must still protect spreadsheet exports because input
  validation is defense in depth, not a substitute for safe output handling.

Design notes:

* Public report and metadata objects are frozen dataclasses. The contained
  pandas DataFrame remains mutable because M10 Cleaning / ETL must transform it.
* Each pipeline stage has one primary responsibility and can be replaced with a
  test double through constructor injection.
* Expected ingestion failures use a typed exception hierarchy rooted at
  IngestionError.
* Raw cell values and filenames are not written to logs. Logs use row and column
  positions plus a short, non-reversible source identifier.
"""
from __future__ import annotations

import csv
import hashlib
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Final, Iterable, Protocol, Sequence
from zipfile import BadZipFile, ZipFile, ZipInfo

from src.logs import get_logger

if TYPE_CHECKING:  # pragma: no cover - imports used only by static type checkers
    import pandas as pd


# --------------------------------------------------------------------------- #
# Contracts
# --------------------------------------------------------------------------- #

REQUIRED_COLUMNS: Final[tuple[str, ...]] = (
    "InvoiceNo",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "UnitPrice",
    "CustomerID",
    "Country",
)

# These prefixes can make a text value execute as a formula when exported to a
# spreadsheet. Pure signed numeric literals such as -1 and +2.5 are allowed.
_FORMULA_PREFIXES: Final[tuple[str, ...]] = (
    "=",
    "+",
    "-",
    "@",
    "\t",
    "\r",
    "\n",
    "＝",
    "＋",
    "－",
    "＠",
)

_SIGNED_NUMERIC_RE: Final[re.Pattern[str]] = re.compile(
    r"^[+-](?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$"
)

MAX_FILE_BYTES: Final[int] = 64 * 1024 * 1024  # 64 MiB
MAX_ROWS: Final[int] = 2_000_000
MAX_XLSX_MEMBERS: Final[int] = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES: Final[int] = 512 * 1024 * 1024  # 512 MiB
MAX_XLSX_COMPRESSION_RATIO: Final[float] = 250.0

_ALLOWED_EXTENSIONS: Final[frozenset[str]] = frozenset({".csv", ".xlsx"})
_XLSX_MAGIC: Final[bytes] = b"PK\x03\x04"
_CSV_SAMPLE_BYTES: Final[int] = 8 * 1024

_XLSX_REQUIRED_MEMBERS: Final[frozenset[str]] = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
    }
)

_LOG = get_logger(__name__)


class FileKind(Enum):
    """Supported input file kinds."""

    CSV = "csv"
    XLSX = "xlsx"


class IssueSeverity(Enum):
    """Severity of one validation finding."""

    ERROR = "error"
    WARNING = "warning"


@dataclass(frozen=True, slots=True)
class ValidationIssue:
    """One finding in the data-quality report."""

    code: str
    message: str
    severity: IssueSeverity
    row: int | None = None
    column: str | None = None


@dataclass(frozen=True, slots=True)
class DataQualityReport:
    """Aggregate result of the validation stage."""

    total_rows: int
    accepted_rows: int
    issues: tuple[ValidationIssue, ...] = field(default_factory=tuple)

    @property
    def has_errors(self) -> bool:
        return any(issue.severity is IssueSeverity.ERROR for issue in self.issues)

    @property
    def error_count(self) -> int:
        return sum(
            1 for issue in self.issues if issue.severity is IssueSeverity.ERROR
        )

    @property
    def warning_count(self) -> int:
        return sum(
            1 for issue in self.issues if issue.severity is IssueSeverity.WARNING
        )


@dataclass(frozen=True, slots=True)
class IngestionResult:
    """Validated frame plus immutable ingestion metadata.

    The dataclass fields cannot be rebound. The DataFrame itself intentionally
    remains mutable for the cleaning stage.
    """

    frame: "pd.DataFrame"
    report: DataQualityReport
    source: Path
    kind: FileKind


# --------------------------------------------------------------------------- #
# Typed exception hierarchy
# --------------------------------------------------------------------------- #


class IngestionError(ValueError):
    """Base class for expected ingestion failures."""


class FileAccessError(IngestionError):
    """The source cannot be found, opened, or read."""


class UnsupportedFileTypeError(IngestionError):
    """The filename extension is not supported."""


class FileTooLargeError(IngestionError):
    """The source exceeds a configured byte, row, or archive safety limit."""


class MimeMismatchError(IngestionError):
    """The extension and detected container type disagree."""


class FileParseError(IngestionError):
    """A supported source cannot be parsed safely."""


class MacroContentError(IngestionError):
    """An XLSX archive contains a macro or unsupported binary part."""


class SchemaViolationError(IngestionError):
    """The input columns do not match the required schema."""


class FormulaInjectionError(IngestionError):
    """A workbook formula or formula-triggering text value was found."""


# --------------------------------------------------------------------------- #
# Stage protocols
# --------------------------------------------------------------------------- #


class _FrameReader(Protocol):
    """Reads at most max_rows + 1 rows so the caller can detect overflow."""

    def read(self, source: Path, max_rows: int) -> "pd.DataFrame": ...


class _SourceInspector(Protocol):
    """Checks the raw source and returns its unmangled header."""

    def inspect(self, source: Path, kind: FileKind) -> tuple[str, ...]: ...


# --------------------------------------------------------------------------- #
# Stage 1: file-type detection
# --------------------------------------------------------------------------- #


class FileTypeDetector:
    """Resolves FileKind from the extension and a small content sample."""

    def detect(self, source: Path) -> FileKind:
        self._require_regular_file(source)

        extension = source.suffix.lower()
        if extension not in _ALLOWED_EXTENSIONS:
            raise UnsupportedFileTypeError(
                f"Unsupported extension {extension!r}; expected one of "
                f"{sorted(_ALLOWED_EXTENSIONS)}"
            )

        try:
            with source.open("rb") as handle:
                sample = handle.read(_CSV_SAMPLE_BYTES)
        except OSError as exc:
            raise FileAccessError(f"Unable to read source file: {exc}") from exc

        if extension == ".xlsx":
            if not sample.startswith(_XLSX_MAGIC):
                raise MimeMismatchError(
                    "Extension is .xlsx but the source is not a ZIP container"
                )
            return FileKind.XLSX

        if sample.startswith(_XLSX_MAGIC):
            raise MimeMismatchError(
                "Extension is .csv but the source is a ZIP container"
            )
        if b"\x00" in sample:
            raise MimeMismatchError(
                "Extension is .csv but the source contains binary NUL bytes"
            )
        return FileKind.CSV

    @staticmethod
    def _require_regular_file(source: Path) -> None:
        try:
            if not source.exists():
                raise FileAccessError(f"File not found: {source}")
            if not source.is_file():
                raise FileAccessError(f"Source is not a regular file: {source}")
        except OSError as exc:
            raise FileAccessError(f"Unable to inspect source file: {exc}") from exc


# --------------------------------------------------------------------------- #
# Stage 2: raw source inspection
# --------------------------------------------------------------------------- #


class SourceInspector:
    """Validates raw headers and XLSX package content before pandas parsing."""

    def __init__(
        self,
        *,
        max_xlsx_members: int = MAX_XLSX_MEMBERS,
        max_xlsx_uncompressed_bytes: int = MAX_XLSX_UNCOMPRESSED_BYTES,
        max_xlsx_compression_ratio: float = MAX_XLSX_COMPRESSION_RATIO,
    ) -> None:
        self._max_xlsx_members = max_xlsx_members
        self._max_xlsx_uncompressed_bytes = max_xlsx_uncompressed_bytes
        self._max_xlsx_compression_ratio = max_xlsx_compression_ratio

    def inspect(self, source: Path, kind: FileKind) -> tuple[str, ...]:
        if kind is FileKind.CSV:
            return self._read_csv_header(source)
        return self._inspect_xlsx(source)

    def _read_csv_header(self, source: Path) -> tuple[str, ...]:
        try:
            return self._read_csv_header_with_encoding(source, "utf-8-sig")
        except UnicodeDecodeError:
            _LOG.info("csv header utf-8 decode failed; retrying latin-1")
            try:
                return self._read_csv_header_with_encoding(source, "latin-1")
            except (csv.Error, OSError, UnicodeError) as exc:
                raise FileParseError(f"Unable to read CSV header: {exc}") from exc
        except (csv.Error, OSError) as exc:
            raise FileParseError(f"Unable to read CSV header: {exc}") from exc

    @staticmethod
    def _read_csv_header_with_encoding(
        source: Path, encoding: str
    ) -> tuple[str, ...]:
        with source.open("r", encoding=encoding, newline="") as handle:
            reader = csv.reader(handle)
            try:
                header = next(reader)
            except StopIteration as exc:
                raise FileParseError("CSV file is empty") from exc

        if not header:
            raise FileParseError("CSV header is empty")
        return tuple(header)

    def _inspect_xlsx(self, source: Path) -> tuple[str, ...]:
        self._inspect_xlsx_archive(source)
        return self._read_xlsx_header(source)

    def _inspect_xlsx_archive(self, source: Path) -> None:
        try:
            with ZipFile(source) as archive:
                infos = archive.infolist()
                names = {info.filename for info in infos}

                missing_members = _XLSX_REQUIRED_MEMBERS.difference(names)
                if missing_members:
                    raise MimeMismatchError(
                        "ZIP container is not a valid XLSX package; missing "
                        f"members: {sorted(missing_members)}"
                    )

                self._enforce_archive_limits(infos)
                self._reject_binary_parts(infos)
                self._reject_worksheet_formulas(archive, infos)
        except BadZipFile as exc:
            raise MimeMismatchError(f"Not a valid XLSX archive: {exc}") from exc
        except OSError as exc:
            raise FileAccessError(f"Unable to inspect XLSX archive: {exc}") from exc

    def _enforce_archive_limits(self, infos: Sequence[ZipInfo]) -> None:
        if len(infos) > self._max_xlsx_members:
            raise FileTooLargeError(
                f"XLSX archive has {len(infos):,} members; maximum is "
                f"{self._max_xlsx_members:,}"
            )

        total_uncompressed = 0
        total_compressed = 0
        for info in infos:
            if info.flag_bits & 0x1:
                raise FileParseError("Encrypted XLSX archive members are not allowed")
            total_uncompressed += info.file_size
            total_compressed += info.compress_size

        if total_uncompressed > self._max_xlsx_uncompressed_bytes:
            raise FileTooLargeError(
                f"XLSX expands to {total_uncompressed:,} bytes; maximum is "
                f"{self._max_xlsx_uncompressed_bytes:,}"
            )

        if total_uncompressed > 0:
            if total_compressed == 0:
                raise FileTooLargeError(
                    "XLSX archive has a suspicious compression ratio"
                )
            ratio = total_uncompressed / total_compressed
            if ratio > self._max_xlsx_compression_ratio:
                raise FileTooLargeError(
                    f"XLSX compression ratio is {ratio:.1f}; maximum is "
                    f"{self._max_xlsx_compression_ratio:.1f}"
                )

    @staticmethod
    def _reject_binary_parts(infos: Sequence[ZipInfo]) -> None:
        for info in infos:
            member = info.filename
            lower = member.lower()
            if "vbaproject.bin" in lower or lower.endswith(".bin"):
                raise MacroContentError(
                    f"Workbook contains a macro or binary part: {member}"
                )

    @staticmethod
    def _reject_worksheet_formulas(
        archive: ZipFile, infos: Sequence[ZipInfo]
    ) -> None:
        worksheet_members = [
            info.filename
            for info in infos
            if info.filename.startswith("xl/worksheets/")
            and info.filename.lower().endswith(".xml")
        ]

        for member in worksheet_members:
            try:
                with archive.open(member) as stream:
                    for _, element in ET.iterparse(stream, events=("end",)):
                        local_name = element.tag.rsplit("}", 1)[-1]
                        if local_name == "f":
                            raise FormulaInjectionError(
                                "Workbook contains a formula in worksheet part "
                                f"{member}"
                            )
                        element.clear()
            except ET.ParseError as exc:
                raise FileParseError(
                    f"Invalid worksheet XML in XLSX member {member}: {exc}"
                ) from exc

    @staticmethod
    def _read_xlsx_header(source: Path) -> tuple[str, ...]:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                source,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            try:
                if not workbook.worksheets:
                    raise FileParseError("XLSX workbook contains no worksheets")

                worksheet = workbook.worksheets[0]
                first_row = next(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=1,
                        values_only=True,
                    ),
                    None,
                )
                if first_row is None:
                    raise FileParseError("XLSX first worksheet is empty")

                return tuple("" if value is None else str(value) for value in first_row)
            finally:
                workbook.close()
        except IngestionError:
            raise
        except Exception as exc:
            raise FileParseError(
                f"Unable to read XLSX header: {type(exc).__name__}: {exc}"
            ) from exc


# --------------------------------------------------------------------------- #
# Stage 3: bounded readers
# --------------------------------------------------------------------------- #


class CsvReader:
    """Reads a bounded CSV sample as strings without numeric coercion."""

    def read(self, source: Path, max_rows: int) -> "pd.DataFrame":
        try:
            return self._read_with_encoding(source, max_rows, "utf-8-sig")
        except UnicodeDecodeError:
            _LOG.info("csv utf-8 decode failed; retrying latin-1")
            try:
                return self._read_with_encoding(source, max_rows, "latin-1")
            except Exception as exc:
                self._raise_csv_parse_error(exc)
        except Exception as exc:
            self._raise_csv_parse_error(exc)

        raise AssertionError("unreachable")

    @staticmethod
    def _read_with_encoding(
        source: Path, max_rows: int, encoding: str
    ) -> "pd.DataFrame":
        import pandas as pd

        return pd.read_csv(
            source,
            dtype=str,
            keep_default_na=False,
            na_values=[""],
            encoding=encoding,
            on_bad_lines="error",
            nrows=max_rows + 1,
        )

    @staticmethod
    def _raise_csv_parse_error(exc: Exception) -> None:
        import pandas as pd

        if isinstance(exc, IngestionError):
            raise exc
        if isinstance(
            exc,
            (
                pd.errors.ParserError,
                pd.errors.EmptyDataError,
                UnicodeError,
                OSError,
                ValueError,
            ),
        ):
            raise FileParseError(f"Unable to parse CSV source: {exc}") from exc
        raise FileParseError(
            f"Unexpected CSV parser failure: {type(exc).__name__}: {exc}"
        ) from exc


class XlsxReader:
    """Reads a bounded XLSX first worksheet after SourceInspector approval."""

    def read(self, source: Path, max_rows: int) -> "pd.DataFrame":
        import pandas as pd

        try:
            return pd.read_excel(
                source,
                engine="openpyxl",
                dtype=str,
                keep_default_na=False,
                na_values=[""],
                nrows=max_rows + 1,
            )
        except IngestionError:
            raise
        except Exception as exc:
            raise FileParseError(
                f"Unable to parse XLSX source: {type(exc).__name__}: {exc}"
            ) from exc


# --------------------------------------------------------------------------- #
# Stage 4: schema validation
# --------------------------------------------------------------------------- #


class SchemaValidator:
    """Requires the exact set of UCI Online Retail columns."""

    def __init__(self, required: Iterable[str] = REQUIRED_COLUMNS) -> None:
        self._required = tuple(required)

    def validate_columns(self, columns: Iterable[object]) -> None:
        actual = tuple(str(column) for column in columns)
        duplicates = self._find_duplicates(actual)
        if duplicates:
            raise SchemaViolationError(
                f"Duplicate column names detected: {duplicates}"
            )

        missing = tuple(column for column in self._required if column not in actual)
        unexpected = tuple(column for column in actual if column not in self._required)

        if missing or unexpected:
            parts: list[str] = []
            if missing:
                parts.append(f"missing={missing}")
            if unexpected:
                parts.append(f"unexpected={unexpected}")
            raise SchemaViolationError(
                "Input columns do not match the required schema: " + "; ".join(parts)
            )

    def validate(self, frame: "pd.DataFrame") -> None:
        self.validate_columns(frame.columns)

    @staticmethod
    def _find_duplicates(columns: Iterable[str]) -> tuple[str, ...]:
        seen: set[str] = set()
        duplicates: list[str] = []

        for column in columns:
            if column in seen and column not in duplicates:
                duplicates.append(column)
            seen.add(column)

        return tuple(duplicates)


# --------------------------------------------------------------------------- #
# Stage 5: cell content inspection
# --------------------------------------------------------------------------- #


class ContentInspector:
    """Screens text cells for spreadsheet formula-injection prefixes.

    Pure signed numeric literals are permitted because negative quantities and
    prices are valid retail data. Other values beginning with a formula prefix
    are rejected or reported, depending on strict mode.
    """

    def __init__(self, *, strict: bool = True) -> None:
        self._strict = strict

    def inspect(self, frame: "pd.DataFrame") -> tuple[ValidationIssue, ...]:
        issues: list[ValidationIssue] = []

        for column_position, column in enumerate(frame.columns, start=1):
            series = frame[column].astype("string")
            normalized = series.str.lstrip(" ")
            has_prefix = normalized.str.startswith(_FORMULA_PREFIXES, na=False)
            is_signed_number = normalized.str.fullmatch(
                _SIGNED_NUMERIC_RE, na=False
            )
            unsafe = has_prefix & ~is_signed_number

            for row_position, is_unsafe in enumerate(
                unsafe.to_numpy(dtype=bool), start=1
            ):
                if not is_unsafe:
                    continue

                issue = ValidationIssue(
                    code="FORMULA_INJECTION",
                    message=(
                        "Cell begins with a formula-triggering character at "
                        f"row {row_position}, column {column!r}"
                    ),
                    severity=IssueSeverity.ERROR,
                    row=row_position,
                    column=str(column),
                )
                _LOG.warning(
                    "formula injection candidate row=%d column_position=%d",
                    row_position,
                    column_position,
                )

                if self._strict:
                    raise FormulaInjectionError(issue.message)
                issues.append(issue)

        return tuple(issues)


# --------------------------------------------------------------------------- #
# Composition root
# --------------------------------------------------------------------------- #


class Ingestor:
    """Composes source checks, bounded parsing, and validation."""

    def __init__(
        self,
        *,
        detector: FileTypeDetector | None = None,
        source_inspector: _SourceInspector | None = None,
        csv_reader: _FrameReader | None = None,
        xlsx_reader: _FrameReader | None = None,
        schema: SchemaValidator | None = None,
        inspector: ContentInspector | None = None,
        max_file_bytes: int = MAX_FILE_BYTES,
        max_rows: int = MAX_ROWS,
    ) -> None:
        if max_file_bytes < 1:
            raise ValueError("max_file_bytes must be positive")
        if max_rows < 1:
            raise ValueError("max_rows must be positive")

        self._detector = detector or FileTypeDetector()
        self._source_inspector = source_inspector or SourceInspector()
        self._csv = csv_reader or CsvReader()
        self._xlsx = xlsx_reader or XlsxReader()
        self._schema = schema or SchemaValidator()
        self._inspector = inspector or ContentInspector(strict=True)
        self._max_file_bytes = max_file_bytes
        self._max_rows = max_rows

    def ingest(self, source: Path | str) -> IngestionResult:
        normalized_source = Path(source)
        kind = self._detector.detect(normalized_source)
        self._enforce_file_size(normalized_source)

        source_id = self._source_id(normalized_source)
        _LOG.info("ingest start source_id=%s kind=%s", source_id, kind.value)

        raw_header = self._source_inspector.inspect(normalized_source, kind)
        self._schema.validate_columns(raw_header)

        reader = self._csv if kind is FileKind.CSV else self._xlsx
        frame = reader.read(normalized_source, self._max_rows)

        if len(frame) > self._max_rows:
            raise FileTooLargeError(
                f"File contains more than {self._max_rows:,} data rows"
            )

        # Validate again after parsing in case a parser or injected reader
        # transforms the columns unexpectedly.
        self._schema.validate(frame)
        issues = self._inspector.inspect(frame)

        error_rows = {
            issue.row
            for issue in issues
            if issue.severity is IssueSeverity.ERROR and issue.row is not None
        }
        total_rows = int(len(frame))
        report = DataQualityReport(
            total_rows=total_rows,
            accepted_rows=total_rows - len(error_rows),
            issues=issues,
        )

        _LOG.info(
            "ingest ok source_id=%s rows=%d accepted=%d errors=%d warnings=%d",
            source_id,
            report.total_rows,
            report.accepted_rows,
            report.error_count,
            report.warning_count,
        )
        return IngestionResult(
            frame=frame,
            report=report,
            source=normalized_source,
            kind=kind,
        )

    def _enforce_file_size(self, source: Path) -> None:
        try:
            size = source.stat().st_size
        except OSError as exc:
            raise FileAccessError(f"Unable to determine source size: {exc}") from exc

        if size > self._max_file_bytes:
            raise FileTooLargeError(
                f"File is {size:,} bytes; maximum is {self._max_file_bytes:,}"
            )

    @staticmethod
    def _source_id(source: Path) -> str:
        digest = hashlib.sha256(source.name.encode("utf-8", errors="replace"))
        return digest.hexdigest()[:12]


# --------------------------------------------------------------------------- #
# Module-level convenience API
# --------------------------------------------------------------------------- #


def ingest(source: Path | str) -> "pd.DataFrame":
    """Load a UCI-schema retail file into a validated DataFrame.

    This convenience function returns only the validated DataFrame.
    Use Ingestor.ingest() when the data-quality report, source path,
    or detected file type is also needed.
    """

    return Ingestor().ingest(source).frame


__all__ = [
    # Value objects
    "REQUIRED_COLUMNS",
    "FileKind",
    "IssueSeverity",
    "ValidationIssue",
    "DataQualityReport",
    "IngestionResult",
    # Exceptions
    "IngestionError",
    "FileAccessError",
    "UnsupportedFileTypeError",
    "FileTooLargeError",
    "MimeMismatchError",
    "FileParseError",
    "MacroContentError",
    "SchemaViolationError",
    "FormulaInjectionError",
    # Stages
    "FileTypeDetector",
    "SourceInspector",
    "CsvReader",
    "XlsxReader",
    "SchemaValidator",
    "ContentInspector",
    # Composition
    "Ingestor",
    "ingest",
]
